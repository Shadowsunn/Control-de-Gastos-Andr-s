import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from datetime import date
import backend
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False

# INICIALIZACIÓN

backend.inicializar_db()

root = tk.Tk()
root.title("Control de Gastos.")
root.geometry("950x650")
root.resizable(False, False)


# NOTEBOOK (TABS)

notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True, padx=10, pady=10)

tab_hoy       = tk.Frame(notebook, padx=15, pady=15)
tab_mensual   = tk.Frame(notebook, padx=15, pady=15)
tab_categoria = tk.Frame(notebook, padx=15, pady=15)

notebook.add(tab_hoy,       text="  📅 Hoy  ")
notebook.add(tab_mensual,   text="  📊 Mensual  ")
notebook.add(tab_categoria, text="  🏷️ Categorías  ")

# TAB 1 HOY

frame_form = tk.LabelFrame(tab_hoy, text="Registrar nuevo gasto", font=("Arial", 11, "bold"), padx=10, pady=10)
frame_form.pack(fill="x", pady=(0, 10))

tk.Label(frame_form, text="Monto ($):",    font=("Arial", 11)).grid(row=0, column=0, sticky="e", padx=5, pady=4)
entry_monto = tk.Entry(frame_form, width=28)
entry_monto.grid(row=0, column=1, padx=5, pady=4)

tk.Label(frame_form, text="Categoría:",   font=("Arial", 11)).grid(row=1, column=0, sticky="e", padx=5, pady=4)
combo_categoria = ttk.Combobox(frame_form, width=26, values=backend.CATEGORIAS_DEFAULT)
combo_categoria.grid(row=1, column=1, padx=5, pady=4)
combo_categoria.current(0)

tk.Label(frame_form, text="Descripción:", font=("Arial", 11)).grid(row=2, column=0, sticky="e", padx=5, pady=4)
entry_descripcion = tk.Entry(frame_form, width=28)
entry_descripcion.grid(row=2, column=1, padx=5, pady=4)

tk.Label(frame_form, text="Tipo:",        font=("Arial", 11)).grid(row=3, column=0, sticky="e", padx=5, pady=4)
combo_tipo = ttk.Combobox(frame_form, width=26, state="readonly", values=["Personal", "Negocio"])
combo_tipo.grid(row=3, column=1, padx=5, pady=4)
combo_tipo.current(0)

tk.Button(
    frame_form, text="✅ Registrar Gasto",
    font=("Arial", 11, "bold"), bg="#28a745", fg="white",
    command=lambda: registrar_gasto()
).grid(row=4, column=0, columnspan=2, pady=12)

frame_tabla_hoy = tk.LabelFrame(tab_hoy, text="Gastos de hoy", font=("Arial", 11, "bold"), padx=5, pady=5)
frame_tabla_hoy.pack(fill="both", expand=True)

tabla_hoy = ttk.Treeview(
    frame_tabla_hoy,
    columns=("fecha", "descripcion", "categoria", "monto", "tipo"),
    show="headings", height=8
)
tabla_hoy.heading("fecha",       text="Fecha")
tabla_hoy.heading("descripcion", text="Descripción")
tabla_hoy.heading("categoria",   text="Categoría")
tabla_hoy.heading("monto",       text="Monto")
tabla_hoy.heading("tipo",        text="Tipo")

tabla_hoy.column("fecha",       width=90,  anchor="center")
tabla_hoy.column("descripcion", width=240)
tabla_hoy.column("categoria",   width=150)
tabla_hoy.column("monto",       width=110, anchor="e")
tabla_hoy.column("tipo",        width=90,  anchor="center")

scroll_hoy = ttk.Scrollbar(frame_tabla_hoy, orient="vertical", command=tabla_hoy.yview)
tabla_hoy.configure(yscrollcommand=scroll_hoy.set)
tabla_hoy.pack(side="left", fill="both", expand=True)
scroll_hoy.pack(side="right", fill="y")

frame_acciones = tk.Frame(tab_hoy)
frame_acciones.pack(fill="x", pady=8)

label_total_hoy = tk.Label(frame_acciones, text="Total del día: $0", font=("Arial", 13, "bold"), fg="#0056b3")
label_total_hoy.pack(side="left", padx=5)

tk.Button(
    frame_acciones, text="🗑️ Eliminar seleccionado",
    font=("Arial", 10, "bold"), bg="#dc3545", fg="white",
    command=lambda: eliminar_gasto_seleccionado()
).pack(side="right", padx=5)

tk.Button(
    frame_acciones, text="✏️ Editar seleccionado",
    font=("Arial", 10, "bold"), bg="#fd7e14", fg="white",
    command=lambda: abrir_ventana_editar()
).pack(side="right", padx=5)

# TAB 2 MENSUAL

frame_selector_mes = tk.Frame(tab_mensual)
frame_selector_mes.pack(fill="x", pady=(0, 10))

tk.Label(frame_selector_mes, text="Mes:", font=("Arial", 11)).pack(side="left", padx=(0, 4))
combo_mes = ttk.Combobox(frame_selector_mes, width=12, state="readonly", values=[
    "1 - Enero", "2 - Febrero", "3 - Marzo", "4 - Abril",
    "5 - Mayo", "6 - Junio", "7 - Julio", "8 - Agosto",
    "9 - Septiembre", "10 - Octubre", "11 - Noviembre", "12 - Diciembre"
])
combo_mes.pack(side="left", padx=(0, 15))
combo_mes.current(date.today().month - 1)

tk.Label(frame_selector_mes, text="Año:", font=("Arial", 11)).pack(side="left", padx=(0, 4))
combo_año = ttk.Combobox(frame_selector_mes, width=8, state="readonly",
                          values=[str(y) for y in range(2024, date.today().year + 2)])
combo_año.pack(side="left", padx=(0, 15))
combo_año.set(str(date.today().year))

tk.Button(
    frame_selector_mes, text="🔍 Buscar",
    font=("Arial", 10, "bold"), bg="#007bff", fg="white",
    command=lambda: cargar_gastos_mes()
).pack(side="left")

frame_tabla_mes = tk.LabelFrame(tab_mensual, text="Gastos del mes", font=("Arial", 11, "bold"), padx=5, pady=5)
frame_tabla_mes.pack(fill="both", expand=True)

tabla_mes = ttk.Treeview(
    frame_tabla_mes,
    columns=("fecha", "descripcion", "categoria", "monto", "tipo"),
    show="headings", height=12
)
tabla_mes.heading("fecha",       text="Fecha")
tabla_mes.heading("descripcion", text="Descripción")
tabla_mes.heading("categoria",   text="Categoría")
tabla_mes.heading("monto",       text="Monto")
tabla_mes.heading("tipo",        text="Tipo")

tabla_mes.column("fecha",       width=90,  anchor="center")
tabla_mes.column("descripcion", width=240)
tabla_mes.column("categoria",   width=150)
tabla_mes.column("monto",       width=110, anchor="e")
tabla_mes.column("tipo",        width=90,  anchor="center")

scroll_mes = ttk.Scrollbar(frame_tabla_mes, orient="vertical", command=tabla_mes.yview)
tabla_mes.configure(yscrollcommand=scroll_mes.set)
tabla_mes.pack(side="left", fill="both", expand=True)
scroll_mes.pack(side="right", fill="y")

frame_resumen = tk.Frame(tab_mensual)
frame_resumen.pack(fill="x", pady=8)

label_total_mes      = tk.Label(frame_resumen, text="Total del mes: $0", font=("Arial", 12, "bold"), fg="#0056b3")
label_total_personal = tk.Label(frame_resumen, text="Personal: $0",      font=("Arial", 11), fg="#6f42c1")
label_total_negocio  = tk.Label(frame_resumen, text="Negocio: $0",       font=("Arial", 11), fg="#20c997")

label_total_mes.pack(side="left", padx=10)
label_total_personal.pack(side="left", padx=10)
label_total_negocio.pack(side="left", padx=10)

tk.Button(
    frame_resumen, text="📥 Exportar",
    font=("Arial", 10, "bold"), bg="#6f42c1", fg="white",
    command=lambda: exportar_mes()
).pack(side="right", padx=10)

# CATEGORIAS

frame_selector_cat = tk.Frame(tab_categoria)
frame_selector_cat.pack(fill="x", pady=(0, 10))

tk.Label(frame_selector_cat, text="Mes:", font=("Arial", 11)).pack(side="left", padx=(0, 4))
combo_mes_cat = ttk.Combobox(frame_selector_cat, width=12, state="readonly", values=[
    "1 - Enero", "2 - Febrero", "3 - Marzo", "4 - Abril",
    "5 - Mayo", "6 - Junio", "7 - Julio", "8 - Agosto",
    "9 - Septiembre", "10 - Octubre", "11 - Noviembre", "12 - Diciembre"
])
combo_mes_cat.pack(side="left", padx=(0, 15))
combo_mes_cat.current(date.today().month - 1)

tk.Label(frame_selector_cat, text="Año:", font=("Arial", 11)).pack(side="left", padx=(0, 4))
combo_año_cat = ttk.Combobox(frame_selector_cat, width=8, state="readonly",
                              values=[str(y) for y in range(2024, date.today().year + 2)])
combo_año_cat.pack(side="left", padx=(0, 15))
combo_año_cat.set(str(date.today().year))

tk.Button(
    frame_selector_cat, text="🔍 Buscar",
    font=("Arial", 10, "bold"), bg="#007bff", fg="white",
    command=lambda: cargar_categorias()
).pack(side="left")

frame_tabla_cat = tk.LabelFrame(tab_categoria, text="Gasto por categoría", font=("Arial", 11, "bold"), padx=5, pady=5)
frame_tabla_cat.pack(fill="both", expand=True)

tabla_cat = ttk.Treeview(
    frame_tabla_cat,
    columns=("categoria", "total"),
    show="headings", height=14
)
tabla_cat.heading("categoria", text="Categoría")
tabla_cat.heading("total",     text="Total gastado")
tabla_cat.column("categoria",  width=350)
tabla_cat.column("total",      width=200, anchor="e")

scroll_cat = ttk.Scrollbar(frame_tabla_cat, orient="vertical", command=tabla_cat.yview)
tabla_cat.configure(yscrollcommand=scroll_cat.set)
tabla_cat.pack(side="left", fill="both", expand=True)
scroll_cat.pack(side="right", fill="y")

label_total_cat = tk.Label(tab_categoria, text="Total: $0", font=("Arial", 12, "bold"), fg="#0056b3")
label_total_cat.pack(pady=8)

canvas_grafico = tk.Canvas(tab_categoria, width=700, height=80, bg="white", highlightthickness=1, highlightbackground="#cccccc")
canvas_grafico.pack(pady=(0, 10))
# FUNCIONES PRINCIPALES

def registrar_gasto():
    exito, mensaje = backend.agregar_gasto(
        entry_descripcion.get(),
        combo_categoria.get(),
        entry_monto.get(),
        combo_tipo.get()
    )
    if exito:
        messagebox.showinfo("Éxito", mensaje)
        entry_monto.delete(0, tk.END)
        entry_descripcion.delete(0, tk.END)
        combo_categoria.current(0)
        combo_tipo.current(0)
        cargar_gastos_hoy()
    else:
        messagebox.showerror("Error", mensaje)

def al_cambiar_tab(event):
    tab_actual = notebook.index(notebook.select())
    if tab_actual == 1:  
        cargar_gastos_mes()
    elif tab_actual == 2:
        cargar_categorias()

notebook.bind("<<NotebookTabChanged>>", al_cambiar_tab)

def cargar_gastos_hoy():
    for item in tabla_hoy.get_children():
        tabla_hoy.delete(item)

    exito, resultado = backend.obtener_gastos_dia()
    total_dia = 0

    if exito:
        for fila in resultado["data"]:
            id_gasto, fecha, descripcion, categoria, monto, tipo = fila
            total_dia += monto
            tabla_hoy.insert("", "end",
                values=(fecha, descripcion, categoria, f"${monto:,.0f}", tipo),
                tags=(str(id_gasto),)
            )
    else:
        messagebox.showerror("Error", f"No se pudieron cargar los gastos: {resultado}")

    label_total_hoy.config(text=f"Total del día: ${total_dia:,.0f} ({resultado['count'] if exito else 0} gastos)")


def eliminar_gasto_seleccionado():
    seleccion = tabla_hoy.selection()
    if not seleccion:
        messagebox.showwarning("Advertencia", "Selecciona un gasto para eliminar.")
        return

    if not messagebox.askyesno("Confirmar", "¿Estás seguro de que quieres eliminar este gasto?"):
        return

    id_gasto = int(tabla_hoy.item(seleccion[0], "tags")[0])
    exito, mensaje = backend.eliminar_gasto(id_gasto)
    if exito:
        cargar_gastos_hoy()
    else:
        messagebox.showerror("Error", mensaje)


def abrir_ventana_editar():
    seleccion = tabla_hoy.selection()
    if not seleccion:
        messagebox.showwarning("Advertencia", "Selecciona un gasto para editar.")
        return

    item     = seleccion[0]
    id_gasto = int(tabla_hoy.item(item, "tags")[0])
    valores  = tabla_hoy.item(item, "values")

    ventana = tk.Toplevel(root)
    ventana.title("Editar Gasto")
    ventana.geometry("380x320")
    ventana.resizable(False, False)
    ventana.grab_set()

    tk.Label(ventana, text="Editar Gasto", font=("Arial", 13, "bold")).pack(pady=(15, 10))
    frame = tk.Frame(ventana, padx=15, pady=5)
    frame.pack()

    tk.Label(frame, text="Fecha (YYYY-MM-DD):", font=("Arial", 10)).grid(row=0, column=0, sticky="e", pady=4)
    e_fecha = tk.Entry(frame, width=22)
    e_fecha.insert(0, valores[0])
    e_fecha.grid(row=0, column=1, pady=4, padx=5)

    tk.Label(frame, text="Descripción:", font=("Arial", 10)).grid(row=1, column=0, sticky="e", pady=4)
    e_desc = tk.Entry(frame, width=22)
    e_desc.insert(0, valores[1])
    e_desc.grid(row=1, column=1, pady=4, padx=5)

    tk.Label(frame, text="Categoría:", font=("Arial", 10)).grid(row=2, column=0, sticky="e", pady=4)
    e_cat = ttk.Combobox(frame, width=20, values=backend.CATEGORIAS_DEFAULT)
    e_cat.set(valores[2])
    e_cat.grid(row=2, column=1, pady=4, padx=5)

    monto_limpio = valores[3].replace("$", "").replace(".", "").replace(",", "")
    tk.Label(frame, text="Monto ($):", font=("Arial", 10)).grid(row=3, column=0, sticky="e", pady=4)
    e_monto = tk.Entry(frame, width=22)
    e_monto.insert(0, monto_limpio)
    e_monto.grid(row=3, column=1, pady=4, padx=5)

    tk.Label(frame, text="Tipo:", font=("Arial", 10)).grid(row=4, column=0, sticky="e", pady=4)
    e_tipo = ttk.Combobox(frame, width=20, state="readonly", values=["Personal", "Negocio"])
    e_tipo.set(valores[4])
    e_tipo.grid(row=4, column=1, pady=4, padx=5)

    def guardar_edicion():
        exito, mensaje = backend.editar_gasto(
            id_gasto,
            e_fecha.get(), e_cat.get(), e_desc.get(), e_monto.get(), e_tipo.get()
        )
        if exito:
            messagebox.showinfo("Éxito", mensaje, parent=ventana)
            ventana.destroy()
            cargar_gastos_hoy()
        else:
            messagebox.showerror("Error", mensaje, parent=ventana)

    tk.Button(
        ventana, text="💾 Guardar cambios",
        font=("Arial", 11, "bold"), bg="#28a745", fg="white",
        command=guardar_edicion
    ).pack(pady=15)


def cargar_gastos_mes():
    for item in tabla_mes.get_children():
        tabla_mes.delete(item)

    mes = int(combo_mes.get().split(" ")[0])
    año = int(combo_año.get())

    exito, resultado = backend.obtener_gastos_mes(mes, año)
    if not exito:
        messagebox.showerror("Error", resultado)
        return

    total = total_personal = total_negocio = 0
    for fila in resultado["data"]:
        _, fecha, descripcion, categoria, monto, tipo = fila
        total += monto
        if tipo == "Personal":
            total_personal += monto
        else:
            total_negocio += monto
        tabla_mes.insert("", "end",
            values=(fecha, descripcion, categoria, f"${monto:,.0f}", tipo)
        )

    label_total_mes.config(text=f"Total del mes: ${total:,.0f}")
    label_total_personal.config(text=f"Personal: ${total_personal:,.0f}")
    label_total_negocio.config(text=f"Negocio: ${total_negocio:,.0f}")

def exportar_mes():
    mes = int(combo_mes.get().split(" ")[0])
    año = int(combo_año.get())
    nombre_mes = combo_mes.get().split(" - ")[1]

    exito, resultado = backend.obtener_gastos_mes(mes, año)
    if not exito:
        messagebox.showerror("Error", resultado)
        return
    if not resultado["data"]:
        messagebox.showwarning("Sin datos", "No hay gastos registrados en ese mes.")
        return

    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
        initialfile=f"gastos_{nombre_mes}_{año}"
    )
    if not ruta:
        return

    if ruta.endswith(".csv"):
        import csv
        with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Fecha", "Descripción", "Categoría", "Monto", "Tipo"])
            for fila in resultado["data"]:
                _, fecha, descripcion, categoria, monto, tipo = fila
                writer.writerow([fecha, descripcion, categoria, monto, tipo])

    else:
        if not OPENPYXL_DISPONIBLE:
            messagebox.showerror(
                "Librería faltante",
                "Para exportar a Excel instala openpyxl:\n\npip install openpyxl"
            )
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{nombre_mes} {año}"

        encabezados = ["Fecha", "Descripción", "Categoría", "Monto", "Tipo"]
        for col, titulo in enumerate(encabezados, start=1):
            celda = ws.cell(row=1, column=col, value=titulo)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="4A90D9")
            celda.alignment = Alignment(horizontal="center")

        for fila_idx, fila in enumerate(resultado["data"], start=2):
            _, fecha, descripcion, categoria, monto, tipo = fila
            ws.cell(row=fila_idx, column=1, value=fecha)
            ws.cell(row=fila_idx, column=2, value=descripcion)
            ws.cell(row=fila_idx, column=3, value=categoria)
            ws.cell(row=fila_idx, column=4, value=monto)
            ws.cell(row=fila_idx, column=5, value=tipo)

        anchos = [12, 35, 20, 15, 12]
        for col, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = ancho

        wb.save(ruta)

    messagebox.showinfo("Éxito", f"Archivo guardado en:\n{ruta}")

def cargar_categorias():
    for item in tabla_cat.get_children():
        tabla_cat.delete(item)

    mes = int(combo_mes_cat.get().split(" ")[0])
    año = int(combo_año_cat.get())

    exito, resultado = backend.obtener_por_categoria(mes, año)
    if not exito:
        messagebox.showerror("Error", resultado)
        return

    total = 0
    for categoria, monto in resultado["data"]:
        total += monto
        tabla_cat.insert("", "end", values=(categoria, f"${monto:,.0f}"))

    label_total_cat.config(text=f"Total: ${total:,.0f}")
    exito2, detalle = backend.obtener_gastos_mes(mes, año)
    total_personal = total_negocio = 0
    if exito2:
        for fila in detalle["data"]:
            _, _, _, _, monto, tipo = fila
            if tipo == "Personal":
                total_personal += monto
            else:
                total_negocio += monto

    dibujar_grafico(total_personal, total_negocio)

def dibujar_grafico(total_personal, total_negocio):
    canvas_grafico.delete("all")

    total = total_personal + total_negocio
    if total == 0:
        canvas_grafico.create_text(350, 40, text="Sin datos para mostrar", font=("Arial", 11), fill="#999999")
        return
    
    ANCHO_BARRA = 600
    X_INICIO = 50
    Y_INICIO = 20
    Y_FIN = 55

    ancho_personal = int(ANCHO_BARRA * (total_personal / total))
    ancho_negocio  = ANCHO_BARRA - ancho_personal

    # Rectángulo Personal (azul)
    canvas_grafico.create_rectangle(
        X_INICIO, Y_INICIO,
        X_INICIO + ancho_personal, Y_FIN,
        fill="#4a90d9", outline=""
    )

    # Rectángulo Negocio (verde)
    canvas_grafico.create_rectangle(
        X_INICIO + ancho_personal, Y_INICIO,
        X_INICIO + ancho_personal + ancho_negocio, Y_FIN,
        fill="#2ecc71", outline=""
    )

    # Etiquetas con porcentaje
    pct_personal = (total_personal / total) * 100
    pct_negocio  = (total_negocio  / total) * 100

    
    # Leyenda centrada
    canvas_grafico.create_rectangle(155, 61, 171, 73, fill="#4a90d9", outline="")
    canvas_grafico.create_text(240, 67, text=f"Personal: {pct_personal:.1f}%", font=("Arial", 10), fill="#333333")

    canvas_grafico.create_rectangle(390, 61, 406, 73, fill="#2ecc71", outline="")
    canvas_grafico.create_text(475, 67, text=f"Negocio: {pct_negocio:.1f}%", font=("Arial", 10), fill="#333333")

cargar_gastos_hoy()

root.mainloop()